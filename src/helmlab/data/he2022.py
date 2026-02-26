"""He 2022 — 3D-printed spherical color difference dataset.

82 color pairs from 3D-printed objects, 5 color groups
(Grey, Red, Green, Yellow, Blue). Uses CIE 1964 10° observer, D65.

Source: https://zenodo.org/records/6502724
"""

import numpy as np
import openpyxl

from helmlab.config import DATA_DIR

HE2022_FILENAME = "he2022_3d_printed.xlsx"

# D65 white for CIE 1964 10° observer (Y=1 scale)
_D65_10DEG = np.array([0.94811, 1.0, 1.07304])


def _lab_to_xyz(Lab: np.ndarray, white: np.ndarray = _D65_10DEG) -> np.ndarray:
    """Convert CIE L*a*b* (10° observer) to XYZ (Y=1 scale).

    Parameters
    ----------
    Lab : ndarray (..., 3)
    white : ndarray (3,) — reference white XYZ

    Returns
    -------
    ndarray (..., 3) — XYZ values
    """
    L, a, b = Lab[..., 0], Lab[..., 1], Lab[..., 2]
    fy = (L + 16.0) / 116.0
    fx = a / 500.0 + fy
    fz = fy - b / 200.0

    delta = 6.0 / 29.0
    delta3 = delta ** 3

    def f_inv(t):
        return np.where(t > delta, t ** 3, 3 * delta ** 2 * (t - 4.0 / 29.0))

    X = white[0] * f_inv(fx)
    Y = white[1] * f_inv(fy)
    Z = white[2] * f_inv(fz)
    return np.stack([X, Y, Z], axis=-1)


def load_he2022() -> dict:
    """Load He 2022 3D-printed color difference dataset.

    Returns
    -------
    dict with keys:
        - XYZ_1: ndarray (82, 3) — first sample XYZ (Y=1 scale)
        - XYZ_2: ndarray (82, 3) — second sample XYZ (Y=1 scale)
        - DV: ndarray (82,) — visual difference (CLAB-scaled)
        - group: list[str] — color group per pair
    """
    path = DATA_DIR / HE2022_FILENAME
    if not path.exists():
        raise FileNotFoundError(
            f"He 2022 dataset not found at {path}. "
            f"Download from https://zenodo.org/records/6502724"
        )

    wb = openpyxl.load_workbook(path)

    # ── Parse samples (CIELAB 10° observer) ──
    ws_samples = wb["CIELAB of 3D printed samples"]
    groups: dict[str, list[dict]] = {}
    current_group = None
    for row in ws_samples.iter_rows(min_row=4, max_row=ws_samples.max_row, values_only=True):
        if isinstance(row[0], str) and row[1] is not None and isinstance(row[1], str):
            current_group = row[0].strip()
            groups[current_group] = []
            continue
        if current_group and isinstance(row[0], (int, float)) and row[1] is not None:
            groups[current_group].append({
                "L": float(row[1]), "a": float(row[2]), "b": float(row[3]),
            })

    # ── Parse pairs ──
    ws_pairs = wb["Information of 3D sample Pairs"]
    pairs = []
    current_group = None
    for row in ws_pairs.iter_rows(min_row=4, max_row=ws_pairs.max_row, values_only=True):
        if row[0] is not None and isinstance(row[0], str):
            if "Experiment" in row[0]:
                continue
            current_group = row[0].strip()
        if (row[1] is not None and isinstance(row[1], str)
                and "Pair" in str(row[1])
                and all(v is not None for v in [row[2], row[7]])):
            pairs.append({
                "group": current_group,
                "DE_ab": float(row[2]),
                "DV": float(row[7]),  # DV (CLAB)
            })

    # ── Match pairs to samples by DE*ab ──
    def _lab_dist(s1, s2):
        return np.sqrt((s1["L"] - s2["L"]) ** 2
                       + (s1["a"] - s2["a"]) ** 2
                       + (s1["b"] - s2["b"]) ** 2)

    Lab_1_list, Lab_2_list, DV_list, group_list = [], [], [], []

    for p in pairs:
        samples = groups[p["group"]]
        best_i, best_j, best_err = 0, 1, 1e9
        for i in range(len(samples)):
            for j in range(i + 1, len(samples)):
                err = abs(_lab_dist(samples[i], samples[j]) - p["DE_ab"])
                if err < best_err:
                    best_err = err
                    best_i, best_j = i, j

        s1, s2 = samples[best_i], samples[best_j]
        Lab_1_list.append([s1["L"], s1["a"], s1["b"]])
        Lab_2_list.append([s2["L"], s2["a"], s2["b"]])
        DV_list.append(p["DV"])
        group_list.append(p["group"])

    Lab_1 = np.array(Lab_1_list, dtype=np.float64)
    Lab_2 = np.array(Lab_2_list, dtype=np.float64)

    # Convert CIELAB (10° observer) → XYZ
    XYZ_1 = _lab_to_xyz(Lab_1)
    XYZ_2 = _lab_to_xyz(Lab_2)
    DV = np.array(DV_list, dtype=np.float64)

    return {
        "XYZ_1": XYZ_1,
        "XYZ_2": XYZ_2,
        "DV": DV,
        "group": group_list,
    }
